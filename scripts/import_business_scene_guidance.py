from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "ai_paths") not in sys.path:
    sys.path.insert(0, str(ROOT / "ai_paths"))

from app.policies.business_scene_table import (  # noqa: E402
    DEFAULT_BUSINESS_SCENE_STATUS,
    build_keywords,
    business_logic_payload,
    canonical_sales_reply_payload,
    copy_strength_for_sales_talk,
    generated_scene_id,
    hard_constraints_for_family,
    infer_policy_family,
    required_tools_for_family,
    risk_rewrite_payload,
    style_reference_payload,
)


DEFAULT_INPUT = (
    Path.home()
    / "Desktop"
    / "\u8d1d\u989c\u6570\u636e"
    / "ai\u5ba2\u670d\u573a\u666f\u95ee\u9898+\u4e1a\u52a1\u903b\u8f91.xlsx"
)
DEFAULT_OUTPUT = ROOT / "ai_paths" / "app" / "policies" / "scene_guidance_business_draft.jsonl"

HEADER_ALIASES = {
    "stage": ("客户阶段", "阶段"),
    "scene_type": ("场景类型", "场景"),
    "question": ("用户问题", "问题"),
    "business_logic": ("业务应答的逻辑", "业务逻辑", "要求以及禁止事项"),
    "sales_talk": ("咨询回答", "销冠话术", "回答建议"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import business scene table into scene guidance JSONL draft.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="Business scene Excel path.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output JSONL path.")
    parser.add_argument(
        "--status",
        default=DEFAULT_BUSINESS_SCENE_STATUS,
        choices=("draft", "shadow", "active"),
        help="Generated scene status. Use draft for review, active only after business validation.",
    )
    parser.add_argument("--sheet", default="", help="Sheet name. Defaults to first worksheet.")
    parser.add_argument("--limit", type=int, default=0, help="Optional row limit for smoke generation.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    rows = load_rows(input_path, sheet_name=args.sheet)
    if args.limit > 0:
        rows = rows[: args.limit]

    records = [build_scene_record(row, status=args.status, source_file=input_path.name) for row in rows]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")

    family_counts: dict[str, int] = {}
    for record in records:
        family = str(record.get("family") or "")
        family_counts[family] = family_counts.get(family, 0) + 1

    print(f"input={input_path}")
    print(f"output={output_path}")
    print(f"records={len(records)} status={args.status}")
    for family, count in sorted(family_counts.items()):
        print(f"{family}={count}")
    return 0


def load_rows(path: Path, *, sheet_name: str = "") -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    iterator = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []
    headers = [str(value or "").strip() for value in header_row]
    indexes = resolve_header_indexes(headers)

    rows: list[dict[str, Any]] = []
    current_stage = ""
    current_scene_type = ""
    for excel_row_number, row in enumerate(iterator, start=2):
        values = [str(value).strip() if value is not None else "" for value in row]
        stage = value_at(values, indexes.get("stage"))
        scene_type = value_at(values, indexes.get("scene_type"))
        question = value_at(values, indexes.get("question"))
        business_logic = value_at(values, indexes.get("business_logic"))
        sales_talk = value_at(values, indexes.get("sales_talk"))
        if stage:
            current_stage = stage
        if scene_type:
            current_scene_type = scene_type
        if not question:
            continue
        rows.append(
            {
                "row_number": excel_row_number,
                "stage": current_stage or stage,
                "scene_type": current_scene_type or scene_type,
                "question": question,
                "business_logic": business_logic,
                "sales_talk": sales_talk,
            }
        )
    return rows


def build_scene_record(row: dict[str, Any], *, status: str, source_file: str) -> dict[str, Any]:
    stage = str(row.get("stage") or "").strip()
    scene_type = str(row.get("scene_type") or "").strip()
    question = str(row.get("question") or "").strip()
    standard = str(row.get("business_logic") or "").strip()
    sales_talk = str(row.get("sales_talk") or "").strip()
    family = infer_policy_family(stage=stage, scene_type=scene_type, question=question, business_logic=standard)
    scene_id = generated_scene_id(
        row_number=int(row.get("row_number") or 0),
        stage=stage,
        scene_type=scene_type,
        question=question,
        family=family,
    )
    keywords = build_keywords(scene_type=scene_type, question=question, business_logic=standard, family=family)
    reply_goal = standard[:120] if standard else f"按{stage}/{scene_type}业务标准承接客户当前问题。"
    return {
        "scene_id": scene_id,
        "family": family,
        "status": status,
        "stage_scope": [stage] if stage else [],
        "examples": [question],
        "keywords": keywords,
        "reply_goal": reply_goal,
        "hard_constraints": hard_constraints_for_family(family),
        "soft_guidance": [
            "先回答客户当前问题",
            "有 canonical_sales_reply 时优先保持其句式、节奏和关键词",
            "只在风险词、事实词和当前客户信息上做最小改写",
        ],
        "business_logic": business_logic_payload(
            family=family,
            stage=stage,
            scene_type=scene_type,
            standard=standard,
        ),
        "style_reference": style_reference_payload(),
        "canonical_sales_reply": canonical_sales_reply_payload(sales_talk),
        "source_sales_reply": sales_talk,
        "copy_strength": copy_strength_for_sales_talk(sales_talk),
        "risk_rewrite": risk_rewrite_payload(sales_talk),
        "source": {
            "type": "business_scene_table",
            "file": source_file,
            "row_number": row.get("row_number"),
            "scene_type": scene_type,
            "stage": stage,
            "required_tools": required_tools_for_family(family),
        },
    }


def resolve_header_indexes(headers: list[str]) -> dict[str, int]:
    resolved: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(headers):
            if any(alias in header for alias in aliases):
                resolved[key] = index
                break
    missing = {"stage", "scene_type", "question"} - set(resolved)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)} headers={headers!r}")
    return resolved


def value_at(values: list[str], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return values[index].strip()


if __name__ == "__main__":
    raise SystemExit(main())

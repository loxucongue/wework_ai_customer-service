from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEMA_VERSION = "sales_strategy_catalog_v1"
CATEGORY_KEYS = {
    "到店受阻": "visit_blocked",
    "距离卡点": "distance_objection",
    "时间不便": "time_objection",
    "价格卡点": "price_objection",
    "效果卡点": "effect_objection",
    "信任疑虑": "trust_objection",
    "健康限制": "health_constraint",
    "修复卡点": "repair_objection",
    "定金卡点": "deposit_objection",
    "决策犹豫": "decision_hesitation",
    "家人决策": "family_decision",
    "需求不匹配": "need_mismatch",
    "意向项目错误": "wrong_project_intent",
}
TIME_BASES = {
    "客户回复时间": "customer_reply",
    "客户加微时间": "contact_added",
    "固定间隔": "previous_step",
    "当天晚6点": "same_day_18_00",
    "当天晚8点": "same_day_20_00",
    "已排客的前一晚8点": "appointment_previous_day_20_00",
}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalized_text(value: Any) -> str:
    return re.sub(r"[\s，。！？、,.!?;；:：()（）\[\]【】\-—]+", "", _text(value)).lower()


def _stable_id(prefix: str, *parts: Any) -> str:
    normalized = "|".join(_normalized_text(part) for part in parts)
    return f"{prefix}_{hashlib.sha1(normalized.encode('utf-8')).hexdigest()[:12]}"


def _split_tags(value: Any) -> list[str]:
    return [item.strip() for item in re.split(r"[/、,，]+", _text(value)) if item.strip()]


def _split_urls(value: Any) -> list[str]:
    return list(dict.fromkeys(item.strip() for item in re.split(r"[\r\n]+", _text(value)) if item.strip()))


def _risk_contract(*values: Any) -> tuple[list[str], list[str]]:
    text = " ".join(_text(value) for value in values)
    required: set[str] = set()
    risks: set[str] = set()
    patterns = (
        (r"天气预报|放晴|下雨|暴雨|天气", "weather_facts", "dynamic_weather"),
        (r"主任|总监|专家|老师外出|老师回来|老师在店", "operator_facts", "dynamic_operator"),
        (r"赠送|额外送|免费送|礼品|赠品", "offer_facts", "dynamic_offer"),
        (r"闭店|开门|营业|停业|装修", "store_operating_facts", "dynamic_store_status"),
        (r"预留|保留名额|锁定名额|已经登记|已经报名|安排好了", "reservation_facts", "dynamic_reservation"),
        (r"活动|优惠|名额|截止|最后一天", "activity_facts", "dynamic_activity"),
        (r"报销车费|包接送|叫车|车费", "transport_support_facts", "dynamic_transport_promise"),
        (r"付款|定金|预约金|订单|支付", "transaction_facts", "transaction_claim"),
    )
    for pattern, fact, risk in patterns:
        if re.search(pattern, text):
            required.add(fact)
            risks.add(risk)
    if re.search(r"100%|百分之百|根治|包治|保证效果|绝对不会", text, flags=re.IGNORECASE):
        risks.add("prohibited_absolute_or_medical_claim")
    return sorted(required), sorted(risks)


def _compile_strategy(path: Path) -> tuple[list[dict[str, Any]], dict[str, set[str]]]:
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    rows = list(worksheet.iter_rows(values_only=True))[2:]
    strategies: list[dict[str, Any]] = []
    aliases: dict[str, set[str]] = {}
    for source_row, row in enumerate(rows, start=3):
        if not any(value not in (None, "") for value in row):
            continue
        category_name = _text(row[1])
        scenario_name = _text(row[2])
        category_key = CATEGORY_KEYS.get(category_name)
        if not category_key or not scenario_name:
            raise ValueError(f"strategy row {source_row} has unknown category or empty scenario")
        scenario_key = _stable_id("scenario", category_key, scenario_name)
        aliases.setdefault(scenario_key, set()).add(scenario_name)
        steps: list[dict[str, Any]] = []
        for index in range(5):
            offset = 3 + index * 4
            timing_name = _text(row[offset])
            raw_delay = row[offset + 1]
            if timing_name == "—" and _text(raw_delay) == "—":
                continue
            if timing_name in TIME_BASES and _text(raw_delay) in TIME_BASES:
                raw_delay = 0
            if not timing_name or timing_name == "—":
                continue
            trigger_base = TIME_BASES.get(timing_name)
            if not trigger_base:
                raise ValueError(f"strategy row {source_row} step {index + 1} has unsupported timing {timing_name}")
            delay = int(float(raw_delay or 0))
            tactic_tags = _split_tags(row[offset + 2])
            description = _text(row[offset + 3])
            required_facts, risk_flags = _risk_contract(description)
            steps.append(
                {
                    "step_key": f"step_{index + 1}",
                    "step_index": index + 1,
                    "trigger_base": trigger_base,
                    "trigger_label": timing_name,
                    "delay_minutes": max(0, delay),
                    "node_goal": description or "结合最新会话化解当前卡点，并只推进一个合法动作。",
                    "tactic_tags": tactic_tags,
                    "description": description,
                    "pressure": "low" if any(tag in {"共情引导", "关怀回访", "低门槛邀请"} for tag in tactic_tags) else "normal",
                    "required_facts": required_facts,
                    "risk_flags": risk_flags,
                }
            )
        strategy_key = _stable_id("strategy", category_key, scenario_name, source_row)
        strategies.append(
            {
                "strategy_key": strategy_key,
                "category_key": category_key,
                "scenario_keys": [scenario_key],
                "name": scenario_name,
                "version": "2026-08-31",
                "enabled": True,
                "steps": steps,
                "source": {"workbook": path.name, "sheet": worksheet.title, "row": source_row},
            }
        )
    return strategies, aliases


def _compile_content(path: Path, aliases: dict[str, set[str]]) -> list[dict[str, Any]]:
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    rows = list(worksheet.iter_rows(values_only=True))[4:]
    contents: list[dict[str, Any]] = []
    fingerprints: dict[str, str] = {}
    for source_row, row in enumerate(rows, start=5):
        if not any(value not in (None, "") for value in row):
            continue
        source_id = _text(row[0])
        category_name = _text(row[1])
        scenario_name = _text(row[2])
        category_key = CATEGORY_KEYS.get(category_name)
        if not category_key or not scenario_name:
            raise ValueError(f"content row {source_row} has unknown category or empty scenario")
        scenario_key = _stable_id("scenario", category_key, scenario_name)
        aliases.setdefault(scenario_key, set()).add(scenario_name)
        solution_idea = _text(row[3])
        answer = _text(row[5])
        image_urls = _split_urls(row[6])
        video_urls = _split_urls(row[7])
        image_url = image_urls[0] if image_urls else ""
        video_url = video_urls[0] if video_urls else ""
        content_types = [name for name, value in (("text", answer), ("image", image_urls), ("video", video_urls)) if value]
        if not content_types:
            if not solution_idea:
                raise ValueError(f"content row {source_row} has no text, media or solution guidance")
            content_types = ["guidance"]
        required_facts, risk_flags = _risk_contract(solution_idea, answer)
        if not answer and not image_url and not video_url:
            risk_flags = sorted({*risk_flags, "guidance_only"})
        media_fingerprints: list[str] = []
        asset_ids: list[str] = []
        for media_value in [*image_urls, *video_urls]:
            digest = hashlib.sha256(media_value.encode("utf-8")).hexdigest()
            fingerprints.setdefault(digest, _stable_id("asset", digest))
            media_fingerprints.append(digest)
            asset_ids.append(fingerprints[digest])
        contents.append(
            {
                "content_id": _stable_id("content", category_key, scenario_name, source_id, answer, image_url, video_url),
                "category_key": category_key,
                "scenario_keys": [scenario_key],
                "scenario_name": scenario_name,
                "tactic_tag": _text(row[4]),
                "solution_idea": solution_idea,
                "reference_text": answer,
                "image_url": image_url,
                "video_url": video_url,
                "image_urls": image_urls,
                "video_urls": video_urls,
                "content_types": content_types,
                "asset_id": asset_ids[0] if asset_ids else "",
                "asset_ids": asset_ids,
                "asset_fingerprint": media_fingerprints[0] if media_fingerprints else "",
                "asset_fingerprints": media_fingerprints,
                "required_facts": required_facts,
                "risk_flags": risk_flags,
                "version": "2026-08-31",
                "status": "active",
                "source": {"workbook": path.name, "sheet": worksheet.title, "row": source_row, "source_id": source_id},
            }
        )
    return contents


def compile_catalog(strategy_path: Path, content_path: Path) -> dict[str, Any]:
    strategies, aliases = _compile_strategy(strategy_path)
    contents = _compile_content(content_path, aliases)
    categories = [
        {"category_key": key, "name": name, "enabled": True}
        for name, key in CATEGORY_KEYS.items()
    ]
    scenarios = []
    for scenario_key, names in sorted(aliases.items()):
        first_name = sorted(names)[0]
        category_key = next(
            item["category_key"]
            for item in contents
            if scenario_key in item["scenario_keys"]
        ) if any(scenario_key in item["scenario_keys"] for item in contents) else next(
            item["category_key"] for item in strategies if scenario_key in item["scenario_keys"]
        )
        scenarios.append(
            {
                "scenario_key": scenario_key,
                "category_key": category_key,
                "name": first_name,
                "aliases": sorted(names),
                "enabled": True,
            }
        )
    result = {
        "schema_version": SCHEMA_VERSION,
        "catalog_version": "2026-08-31.1",
        "status": "published",
        "runtime_mode": "shadow",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "categories": categories,
        "scenarios": scenarios,
        "strategies": strategies,
        "contents": contents,
    }
    canonical = json.dumps(
        {key: value for key, value in result.items() if key != "generated_at"},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    result["checksum"] = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Compile sales strategy Excel workbooks into runtime JSON")
    parser.add_argument("--strategy", type=Path, required=True)
    parser.add_argument("--content", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    catalog = compile_catalog(args.strategy, args.content)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output),
        "categories": len(catalog["categories"]),
        "scenarios": len(catalog["scenarios"]),
        "strategies": len(catalog["strategies"]),
        "contents": len(catalog["contents"]),
        "checksum": catalog["checksum"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()

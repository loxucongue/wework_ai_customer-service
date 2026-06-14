from __future__ import annotations

import argparse
import asyncio
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import get_settings  # noqa: E402
from app.services.coze_client import CozeClient  # noqa: E402


DEFAULT_XLSX = Path(r"C:\Users\24159\Desktop\贝颜数据\ai客服场景问题+业务逻辑.xlsx")
DEFAULT_OUTPUT_DIR = Path("logs")
KB_NAME = "sales_talk_qa"


@dataclass(frozen=True)
class RecallCase:
    index: int
    customer_stage: str
    scene_type: str
    question: str
    business_logic: str
    sales_script: str


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_header(value: Any) -> str:
    return _cell_text(value).replace(" ", "").replace("\n", "")


def _pick(mapping: dict[str, int], *names: str) -> int | None:
    normalized = {_normalize_header(name): name for name in names}
    for header, column in mapping.items():
        if header in normalized:
            return column
    for header, column in mapping.items():
        if any(name in header for name in normalized):
            return column
    return None


def load_cases(path: Path, *, sheet: str | None = None) -> list[RecallCase]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_index = 0
    header_map: dict[str, int] = {}
    for row_index, row in enumerate(rows[:20]):
        candidate = {_normalize_header(value): index for index, value in enumerate(row) if _normalize_header(value)}
        if _pick(candidate, "用户问题", "问题") is not None:
            header_row_index = row_index
            header_map = candidate
            break
    if not header_map:
        raise ValueError("未找到包含“用户问题/问题”的表头行")

    question_col = _pick(header_map, "用户问题", "问题")
    stage_col = _pick(header_map, "客户阶段", "阶段")
    scene_col = _pick(header_map, "场景类型", "场景")
    logic_col = _pick(header_map, "业务应答逻辑", "回答建议", "业务逻辑")
    script_col = _pick(header_map, "咨询回答（销冠话术）", "销冠话术", "参考话术", "回答")
    if question_col is None:
        raise ValueError("未找到用户问题列")

    cases: list[RecallCase] = []
    for index, row in enumerate(rows[header_row_index + 1 :], start=1):
        question = _cell_text(row[question_col] if question_col < len(row) else "")
        if not question or question in {"问题", "用户问题"}:
            continue
        cases.append(
            RecallCase(
                index=index,
                customer_stage=_cell_text(row[stage_col] if stage_col is not None and stage_col < len(row) else ""),
                scene_type=_cell_text(row[scene_col] if scene_col is not None and scene_col < len(row) else ""),
                question=question,
                business_logic=_cell_text(row[logic_col] if logic_col is not None and logic_col < len(row) else ""),
                sales_script=_cell_text(row[script_col] if script_col is not None and script_col < len(row) else ""),
            )
        )
    return cases


def _similarity(expected: str, actual: str) -> float:
    left = "".join(str(expected or "").split())
    right = "".join(str(actual or "").split())
    if not left or not right:
        return 0.0
    return round(SequenceMatcher(None, left, right).ratio(), 4)


def _excerpt(text: str, limit: int = 180) -> str:
    value = " ".join(str(text or "").split())
    return value[:limit]


async def evaluate_case(client: CozeClient, case: RecallCase, semaphore: asyncio.Semaphore) -> dict[str, Any]:
    async with semaphore:
        started = datetime.now()
        try:
            result = await client.search_kb(KB_NAME, case.question)
            first = result.items[0] if result.items else None
            combined = "\n".join(item.content for item in result.items[:3])
            return {
                "index": case.index,
                "customer_stage": case.customer_stage,
                "scene_type": case.scene_type,
                "question": case.question,
                "business_logic": case.business_logic,
                "sales_script": case.sales_script,
                "kb_name": KB_NAME,
                "hit_count": len(result.items),
                "first_document_id": first.document_id if first else "",
                "first_content": _excerpt(first.content if first else ""),
                "top3_similarity_to_sales_script": _similarity(case.sales_script, combined),
                "status": "ok",
                "elapsed_ms": int((datetime.now() - started).total_seconds() * 1000),
            }
        except Exception as exc:
            return {
                "index": case.index,
                "customer_stage": case.customer_stage,
                "scene_type": case.scene_type,
                "question": case.question,
                "business_logic": case.business_logic,
                "sales_script": case.sales_script,
                "kb_name": KB_NAME,
                "hit_count": 0,
                "first_document_id": "",
                "first_content": "",
                "top3_similarity_to_sales_script": 0.0,
                "status": "error",
                "error": f"{type(exc).__name__}: {exc}",
                "elapsed_ms": int((datetime.now() - started).total_seconds() * 1000),
            }


def write_reports(results: list[dict[str, Any]], output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    jsonl_path = output_dir / f"sales_talk_qa_recall_results_{stamp}.jsonl"
    md_path = output_dir / f"sales_talk_qa_recall_report_{stamp}.md"

    jsonl_path.write_text(
        "\n".join(json.dumps(item, ensure_ascii=False) for item in results) + "\n",
        encoding="utf-8",
    )

    total = len(results)
    hit = sum(1 for item in results if int(item.get("hit_count") or 0) > 0)
    errors = sum(1 for item in results if item.get("status") == "error")
    high_similarity = sum(1 for item in results if float(item.get("top3_similarity_to_sales_script") or 0) >= 0.35)
    lines = [
        "# sales_talk_qa 召回评估报告",
        "",
        f"- 总数：{total}",
        f"- 有命中：{hit}",
        f"- 请求失败：{errors}",
        f"- 与销冠话术相似度 >= 0.35：{high_similarity}",
        "",
        "|序号|客户阶段|场景类型|用户问题|命中数|相似度|首条切片摘要|错误|",
        "|---:|---|---|---|---:|---:|---|---|",
    ]
    for item in results:
        lines.append(
            "|{index}|{customer_stage}|{scene_type}|{question}|{hit_count}|{score}|{first}|{error}|".format(
                index=item.get("index", ""),
                customer_stage=_escape_md(item.get("customer_stage", "")),
                scene_type=_escape_md(item.get("scene_type", "")),
                question=_escape_md(item.get("question", "")),
                hit_count=item.get("hit_count", 0),
                score=item.get("top3_similarity_to_sales_script", 0),
                first=_escape_md(item.get("first_content", "")),
                error=_escape_md(item.get("error", "")),
            )
        )
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return jsonl_path, md_path


def _escape_md(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", " ").strip()


async def main_async() -> None:
    parser = argparse.ArgumentParser(description="Evaluate sales_talk_qa recall with original customer questions.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    cases = load_cases(args.xlsx, sheet=args.sheet)
    if args.limit and args.limit > 0:
        cases = cases[: args.limit]
    client = CozeClient(get_settings())
    semaphore = asyncio.Semaphore(max(1, args.workers))
    try:
        results = await asyncio.gather(*(evaluate_case(client, case, semaphore) for case in cases))
    finally:
        await client.aclose()
    jsonl_path, md_path = write_reports(list(results), args.output_dir)
    print(json.dumps({"total": len(results), "jsonl": str(jsonl_path), "report": str(md_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    asyncio.run(main_async())
